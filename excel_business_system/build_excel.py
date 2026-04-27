"""
Builds FitZoneSystem.xlsx from the original BusinessSystem.xlsx template.

Approach:
- Load the original workbook (preserves cell formatting, formulas, conditional formatting,
  and chart objects).
- Replace ONLY the data values (product names, customer info, supplier names, prices,
  quantities, dates, expense rows) and a few hard-coded labels in Analysis / WhatIf /
  Pivots sheets.  Every formula in the original workbook is intentionally left untouched
  so it re-computes correctly against the new data.

Run:
    python build_excel.py /path/to/original/BusinessSystem.xlsx
        /path/to/output/FitZoneSystem.xlsx
"""

from __future__ import annotations

import datetime as _dt
import sys
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
#  FitZone Sports business definition
# ---------------------------------------------------------------------------
PRODUCTS = [
    # ID, Name, Category, Unit_Price, Stock_Quantity
    (1,  "Treadmill Pro X",        "Gym",       28000,   6),
    (2,  "Garmin GPS Watch",       "Running",    8500,  25),
    (3,  "Yoga Mat Premium",       "Gym",         450, 200),
    (4,  "Trek Mountain Bike",     "Cycling",   18000,   9),
    (5,  "Asics Running Shoes",    "Running",    3200,  60),
    (6,  "Speedo Swim Goggles",    "Swimming",    650, 150),
    (7,  "Adidas Football Ball",   "Football",    800, 120),
    (8,  "Nike Football Cleats",   "Football",   2500,  45),
    (9,  "Arena Swimsuit (Men)",   "Swimming",   1200,   8),
    (10, "Dumbbell Set 20 kg",     "Gym",        1800,  30),
]

CUSTOMERS = [
    # ID, Name, City, Contact
    (1, "Mohamed Elsaeed", "Mansoura",   "01000000001"),
    (2, "Ahmed Hassan",    "Cairo",      "01000000002"),
    (3, "Omar Mostafa",    "Mansoura",   "01000000003"),
    (4, "Yasmin Adel",     "Alexandria", "01000000004"),
    (5, "Karim Said",      "Cairo",      "01000000005"),
    (6, "Nada Tarek",      "Mansoura",   "01000000006"),
]

SUPPLIERS = [
    "Adidas Egypt", "Nike ME", "FitnessPro Imports",
    "Decathlon Egypt", "Garmin Distributor", "Trek Bicycles",
]

# Category list (5 categories — the original used 4)
CATEGORIES = ["Football", "Gym", "Swimming", "Running", "Cycling"]

# Quick lookup: Product_ID -> Unit_Price  (used by the Sales table)
PRICE = {p[0]: p[3] for p in PRODUCTS}


# Sales: 20 transactions — same structure as the original (ID, ProductID,
# CustomerID, Quantity, UnitPrice, Date).
SALES = [
    (101, 1,  1,  2, _dt.date(2026, 1,  5)),
    (102, 3,  2,  4, _dt.date(2026, 1,  8)),
    (103, 4,  3,  1, _dt.date(2026, 1, 12)),
    (104, 2,  1,  1, _dt.date(2026, 2,  3)),
    (105, 6,  4, 50, _dt.date(2026, 2,  9)),
    (106, 8,  5,  2, _dt.date(2026, 2, 18)),
    (107, 7,  2,  6, _dt.date(2026, 3,  1)),
    (108, 5,  3,  1, _dt.date(2026, 3,  7)),
    (109, 10, 6, 10, _dt.date(2026, 3, 15)),
    (110, 1,  4,  1, _dt.date(2026, 3, 22)),
    (111, 9,  5,  3, _dt.date(2026, 4,  4)),
    (112, 3,  1,  2, _dt.date(2026, 4, 11)),
    (113, 4,  2,  1, _dt.date(2026, 4, 19)),
    (114, 6,  3, 25, _dt.date(2026, 5,  2)),
    (115, 8,  6,  1, _dt.date(2026, 5,  9)),
    (116, 7,  4,  3, _dt.date(2026, 5, 18)),
    (117, 1,  5,  3, _dt.date(2026, 6,  1)),
    (118, 2,  6,  2, _dt.date(2026, 6, 12)),
    (119, 10, 1,  5, _dt.date(2026, 6, 20)),
    (120, 5,  2,  1, _dt.date(2026, 6, 28)),
]


# Purchases: 12 supplier orders — same row count as the original.
PURCHASES = [
    # ID, Supplier, Product_ID, Quantity_Purchased, Unit_Cost, Date
    (201, "Adidas Egypt",         7, 300,    500, _dt.date(2025, 12,  1)),
    (202, "Nike ME",               8, 100,  1800, _dt.date(2025, 12,  5)),
    (203, "FitnessPro Imports",    1,  10, 22000, _dt.date(2025, 12, 10)),
    (204, "FitnessPro Imports",   10,  80,  1300, _dt.date(2025, 12, 15)),
    (205, "Decathlon Egypt",       3, 500,   280, _dt.date(2025, 12, 18)),
    (206, "Decathlon Egypt",       6, 400,   400, _dt.date(2025, 12, 22)),
    (207, "Garmin Distributor",    2,  40,  6500, _dt.date(2026,  1,  5)),
    (208, "Trek Bicycles",         4,  15, 14000, _dt.date(2026,  1, 12)),
    (209, "FitnessPro Imports",    5,  80,  2300, _dt.date(2026,  1, 18)),
    (210, "Decathlon Egypt",       9,  30,   850, _dt.date(2026,  2,  2)),
    (211, "Adidas Egypt",          7, 200,   480, _dt.date(2026,  2, 14)),
    (212, "Nike ME",               8,  50,  1850, _dt.date(2026,  3,  3)),
]


EXPENSES = [
    # ID, Type, Amount, Date, Notes, Essential
    (301, "Rent",         12000, _dt.date(2026, 1,  1), "Shop monthly rent",      "Yes"),
    (302, "Salaries",     22000, _dt.date(2026, 1,  5), "Monthly payroll",        "Yes"),
    (303, "Electricity",   3000, _dt.date(2026, 1, 10), "Utility",                "Yes"),
    (304, "Marketing",     3500, _dt.date(2026, 1, 15), "Facebook ads",           "No"),
    (305, "Cleaning",       800, _dt.date(2026, 1, 20), "Cleaning service",       "Yes"),
    (306, "Rent",         12000, _dt.date(2026, 2,  1), "Shop monthly rent",      "Yes"),
    (307, "Salaries",     22000, _dt.date(2026, 2,  5), "Monthly payroll",        "Yes"),
    (308, "Marketing",     5500, _dt.date(2026, 2, 15), "Google ads (high)",      "No"),
    (309, "Maintenance",   2200, _dt.date(2026, 2, 22), "Treadmill servicing",    "Yes"),
    (310, "Cleaning",       800, _dt.date(2026, 2, 25), "Cleaning service",       "Yes"),
    (311, "Rent",         12000, _dt.date(2026, 3,  1), "Shop monthly rent",      "Yes"),
    (312, "Salaries",     22000, _dt.date(2026, 3,  5), "Monthly payroll",        "Yes"),
    (313, "Electricity",   2800, _dt.date(2026, 3, 10), "Utility",                "Yes"),
    (314, "Marketing",     1500, _dt.date(2026, 3, 15), "Print flyers",           "No"),
    (315, "Office",         900, _dt.date(2026, 3, 20), "Stationery",             "No"),
]


# ---------------------------------------------------------------------------
#  Builder
# ---------------------------------------------------------------------------
def populate_product(ws):
    for i, (pid, name, cat, price, stock) in enumerate(PRODUCTS, start=2):
        ws.cell(row=i, column=1, value=pid)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=cat)
        ws.cell(row=i, column=4, value=price)
        ws.cell(row=i, column=5, value=stock)
        # F (Stock_Value), G (Avg_Product_Price), H (Stock_Status) keep their formulas


def populate_customers(ws):
    for i, (cid, name, city, contact) in enumerate(CUSTOMERS, start=2):
        ws.cell(row=i, column=1, value=cid)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=city)
        ws.cell(row=i, column=4, value=contact)
        # E, F, G are formulas (Total_Purchases / Loyalty / Customer_Rank)


def populate_sales(ws):
    for i, (sid, prod_id, cust_id, qty, date) in enumerate(SALES, start=2):
        ws.cell(row=i, column=1, value=sid)
        ws.cell(row=i, column=2, value=prod_id)
        ws.cell(row=i, column=3, value=cust_id)
        ws.cell(row=i, column=4, value=qty)
        ws.cell(row=i, column=5, value=PRICE[prod_id])
        # F (Total_Sale) is a formula
        ws.cell(row=i, column=7, value=date)
        # H..O are formulas


def populate_purchases(ws):
    for i, (puid, supplier, prod_id, qty, cost, date) in enumerate(PURCHASES, start=2):
        ws.cell(row=i, column=1, value=puid)
        ws.cell(row=i, column=2, value=supplier)
        ws.cell(row=i, column=3, value=prod_id)
        ws.cell(row=i, column=4, value=qty)
        ws.cell(row=i, column=5, value=cost)
        # F is formula
        ws.cell(row=i, column=7, value=date)
        # H, I are formulas


def populate_expenses(ws):
    for i, (eid, etype, amount, date, notes, essential) in enumerate(EXPENSES, start=2):
        ws.cell(row=i, column=1, value=eid)
        ws.cell(row=i, column=2, value=etype)
        ws.cell(row=i, column=3, value=amount)
        ws.cell(row=i, column=4, value=date)
        ws.cell(row=i, column=5, value=notes)
        # F (Flag) is formula
        ws.cell(row=i, column=7, value=essential)
        # H (After_Cut_10pct) is formula


def patch_analysis(ws):
    """
    The Analysis sheet has hard-coded supplier names in A4:A9 (with formulas in
    B/C using those names) and product names in B13:B22.  We rewrite those
    labels and refresh the SUMIF / AVERAGEIF formulas so they reference the new
    supplier names.
    """
    for i, supplier in enumerate(SUPPLIERS, start=4):
        ws.cell(row=i, column=1, value=supplier)
        ws.cell(row=i, column=2,
                value=f'=SUMIF(Purchases!$B$2:$B$50,"{supplier}",Purchases!$F$2:$F$50)')
        ws.cell(row=i, column=3,
                value=f'=IFERROR(AVERAGEIF(Purchases!$B$2:$B$50,"{supplier}",'
                      f'Purchases!$D$2:$D$50),0)')

    # Product names in B13..B22  (rows 13..22, product_id 1..10)
    for i, (_pid, pname, *_rest) in enumerate(PRODUCTS, start=13):
        ws.cell(row=i, column=2, value=pname)


def patch_whatif(ws):
    """The WhatIf sheet hard-codes 'Apple Egypt' in scenario 4 (cells A30 / B31)."""
    # The "+10% supplier cost" scenario is built around the largest supplier; we
    # repoint it to FitnessPro Imports (high-value treadmill / dumbbell orders).
    target = "FitnessPro Imports"
    ws["A30"] = f"4) {target} unit cost +10% — Total Purchases impact"
    ws["A31"] = f"Current Total ({target})"
    ws["B31"] = f'=SUMIF(Purchases!$B$2:$B$50,"{target}",Purchases!$F$2:$F$50)'


def patch_pivots(ws):
    """Replace category list (A6..A9, A28..A31) and product names (B14..B23)."""
    # Profit per Category — rows 6..(6+len-1)
    for i, cat in enumerate(CATEGORIES, start=6):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2,
                value=f'=SUMIFS(Sales!$F$2:$F$50,Sales!$I$2:$I$50,"{cat}")')
        ws.cell(row=i, column=3,
                value=f'=SUMIFS(Sales!$N$2:$N$50,Sales!$I$2:$I$50,"{cat}")')
        ws.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)")

    # Total Purchases by Product Category — rows 28..(28+len-1)
    for i, cat in enumerate(CATEGORIES, start=28):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2,
                value=f'=SUMIFS(Purchases!$F$2:$F$50,Purchases!$I$2:$I$50,"{cat}")')

    # Product names in B14..B23
    for i, (_pid, pname, *_rest) in enumerate(PRODUCTS, start=14):
        ws.cell(row=i, column=2, value=pname)


def patch_dashboard(ws):
    """The Dashboard title cell."""
    ws["A1"] = "FitZone Sports — Business Dashboard"


def main(template_path: str, output_path: str):
    wb = load_workbook(template_path)

    populate_product(wb["Product"])
    populate_customers(wb["Customers"])
    populate_sales(wb["Sales"])
    populate_purchases(wb["Purchases"])
    populate_expenses(wb["Expenses"])

    patch_analysis(wb["Analysis"])
    patch_whatif(wb["WhatIf"])
    patch_pivots(wb["Pivots"])
    patch_dashboard(wb["Dashboard"])

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    template = sys.argv[1] if len(sys.argv) > 1 else \
        "../../Data-Mining-Tasks/excel_business_system/BusinessSystem.xlsx"
    output = sys.argv[2] if len(sys.argv) > 2 else "FitZoneSystem.xlsx"
    main(template, output)
